# -*- coding:utf8 -*-


from openpyxl.reader.excel import load_workbook

import os
from model import SalaryAdjustRecord
from commons import MyCalendar


class SalaryAdjustDao(object):

    def getAllSalaryAdjustRecords(self , fn , firstDate):
        wb = load_workbook(os.path.expandvars(fn))
        ws = wb.active
        all_map = {}
        firstDate = MyCalendar.get_1st_of_this_month(firstDate)
        lastDate = MyCalendar.get_last_date(firstDate)
        for row_idx in xrange(2,ws.max_row+1):
            r =SalaryAdjustRecord()
            r.init(ws,row_idx)
            if r.effectDate and r.effectDate <=lastDate:
                all_map[r.workcode] = r
        return all_map
    pass

import sys
reload(sys)
sys.setdefaultencoding('utf-8')

d = SalaryAdjustDao()

d.load(u"调薪表.xlsx")




